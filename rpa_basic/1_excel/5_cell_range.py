from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active


# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"]) # A B C
for i in range(1, 6): # 5개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])


print("------------(1) 컬럼 하나 출력---------")
col_B = ws["B"] # 영어 column만 가지고 오기
for cell in col_B:
    print(cell.value)

print("------------(2) 복수 컬럼 출력---------")
col_range = ws["B:C"] # 영어, 수학 column 함께 가지고 오기
for cols in col_range:
    for cell in cols:
        print(cell.value)

print("------------(3) 로우 하나 출력---------")
row_title = ws[1] # 1번재 row 만 가지고오기
for cell in row_title:
    print(cell.value)

from openpyxl.utils.cell import coordinate_from_string

print("------------(4) coordinate 를 이용하여 컬럼의 인덱스 확인 가능---------")
row_range = ws[2:5] # 2번째 줄에서 5번째 줄까지 가지고 오기. slice와 범위가 조금 다르다.
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ")
        print(xy[0], end="") # A
        print(xy[1], end=" ") # 1
    print()

# 전체 rows
# print(tuple(ws.rows)) # 한 줄씩 가져와서 tuple로 저장
print("------------(5) tuple rows---------")
for row in tuple(ws.rows):
    print(row)

for row in tuple(ws.rows):
    print(row[0].value)

# 전체 columns
# print(tuple(ws.columns)) # 한 열씩 가져와서 tuple로 저장
print("------------(6) tuple columns---------")
for column in tuple(ws.columns):
    print(column[0].value)


print("------------(7) iter_rows---------")
for row in ws.iter_rows(): # 전체 row
    print(row[1].value)

print("------------(8) iter_rows---------")
for cols in ws.iter_cols(): # 전체 columns
    print(cols[1].value)

# 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지
print("------------(9) iter_rows 범위 지정---------")
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[0].value, row[1].value) # 수학, 영어

print("------------(10) iter_cols 범위 지정---------")
for col in ws.iter_cols(min_row=1, max_row=5, min_col=2, max_col=3):
    print(col)



wb.save("sample.xlsx")