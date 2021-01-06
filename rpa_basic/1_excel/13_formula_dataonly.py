from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx")
ws = wb.active

# 수식이 그대로 표시됨
for row in ws.values:
    for cell in row:
        print(cell)


# 수식이 아닌 실제 데이터를 가지고 옴
# evalueate 되지 않은 상태의 데이터는 None이라고 표시 >> 이 경우 파일을 열어서 저장을 해야함
print("----------data_only=True---------")
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)

