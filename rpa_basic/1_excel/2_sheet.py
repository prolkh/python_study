from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet() # 새로운 sheet 생성
ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "3366ff" # RGB 형태로 값 넣어주면 탭 색상 변경

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # e두번째 인덱스에 sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근

print(wb.sheetnames)

# Sheet 복사
new_ws["A1"] = "test" # A1 셀에 test 입력
target = wb.copy_worksheet(new_ws) #  sheet 복사

wb.save("sample.xlsx")
