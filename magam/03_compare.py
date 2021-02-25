from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook("magamsample.xlsx") # sample.xlsx 파일에서 워크북을 불러옴
ws = wb.active # 활성화된 Sheet

# cell 개수를 모를 때
# for x in range(1, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         print(ws.cell(row=x, column=y).value, end = " ") # 1 2 3 4 ..
#     print()

x = 1
y = 1
# while (ws.cell(x, 1).value is not None) and (ws.cell(y, 2).value is not None):
#     # print(ws.cell(x, 1).value, ws.cell(y, 2).value)
#     print(x, y)
#     if(ws.cell(x, 1).value > ws.cell(y, 2).value):
#         ws.move_range(f"A{x}:A{ws.max_row}", rows=1, cols=0)
#     elif(ws.cell(x, 1).value < ws.cell(y, 2).value):
#         ws.move_range(f"B{y}:B{ws.max_row}", rows=1, cols=0)
        
#     x = x + 1
#     y = y + 1


import re

print(re.sub("[a-z|A-Z]", "", "A2:F2", ))

while (ws[f"A{x}"].value is not None) and (ws[f"B{y}"].value is not None):
    # print(ws.cell(x, 1).value, ws.cell(y, 2).value)
    print(ws[f"A{x}"].value, ws[f"B{y}"].value)
    left  = ws[f"A{x}"].value
    right = ws[f"B{y}"].value

    if(left > right):
        ws.move_range(f"A{x}:A{ws.max_row}", rows=1, cols=0)
    elif(left < right):
        ws.move_range(f"B{y}:B{ws.max_row}", rows=1, cols=0)
        
    x = x + 1
    y = y + 1




wb.save("magamsample2.xlsx")


