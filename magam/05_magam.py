import tkinter.ttk as ttk
import tkinter.messagebox as messagebox
from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook # 파일 불러오기
import os
import time
import re


root = Tk()
root.title("magam")
# root.geometry("360x360") # 가로 * 세로

def add_file():
    files = filedialog.askopenfilenames(title="엑셀 파일을 불러옵니다.", \
        filetypes=(("엑셀 파일", "*.xl*;"), ("모든 파일", "*.*")), \
        # initialdir=r"D:\python\python_study\pygame_project\images") # 최초에 사용자가 지정한 경로를 보여줌
        initialdir="C:/") # 최초경로
    
    # 사용자가 선택한 파일 목록
    for file in files:
        list_file.insert(END, file)

def del_file():
    for index in reversed(list_file.curselection()):
        list_file.delete(index)


def start():
    # 파일 목록 확인
    if list_file.size() == 0 :
        messagebox.showwarning("경고", "엑셀 파일을 선택하세요")
        return

    # 엑셀 파일 읽기
    wb = load_workbook(list_file.get(0)) # sample.xlsx 파일에서 워크북을 불러옴
    ws = wb.active # 활성화된 Sheet

    # 데이터 시작 위치
    start = opt_rows_combo.get()
    if start == "1행":
        x = y = 1
    elif start == "2행":
        x = y = 2
    elif start == "3행":
        x = y = 3

    # 데이터 범위
    r1_0 = range1_entry.get().split(":")[0]
    r1_1 = range1_entry.get().split(":")[1]

    r2_0 = range2_entry.get().split(":")[0]
    r2_1 = range2_entry.get().split(":")[1]

    # 비교1, 비교2
    c1_e1 = comp1_entry1.get()
    c1_e2 = comp1_entry2.get()
    c2_e1 = comp2_entry1.get()
    c2_e2 = comp2_entry2.get()


    # 파일 정렬
    while (ws[f"{c1_e1}{x}"].value is not None) and (ws[f"{c1_e2}{y}"].value is not None):
        # print(ws[f"{c1_e1}{x}"].value, ws[f"{c1_e2}{y}"].value)
        left  = ws[f"{c1_e1}{x}"].value
        right = ws[f"{c1_e2}{y}"].value

        left2  = ws[f"{c2_e1}{x}"].value
        right2 = ws[f"{c2_e2}{y}"].value

        if(left2 is None):
            left2 = 0

        if(right2 is None):
            right2 = 0

        if(left > right):
            ws.move_range(f"{r1_0}{x}:{r1_1}{ws.max_row}", rows=1, cols=0)
        elif(left < right):
            ws.move_range(f"{r2_0}{y}:{r2_1}{ws.max_row}", rows=1, cols=0)
        elif(left == right and left2 > right2):
            ws.move_range(f"{r1_0}{x}:{r1_1}{ws.max_row}", rows=1, cols=0)
        elif(left == right and left2 < right2):
            ws.move_range(f"{r2_0}{y}:{r2_1}{ws.max_row}", rows=1, cols=0)

        x = x + 1
        y = y + 1

    # 확장자
    name, ext = os.path.splitext(list_file.get(0))
    # 새로운 엑셀 저장
    file_name = "{}{}{}".format(name, time.strftime("_%Y%m%d_%H%M%S"), ext)
    dest_path = os.path.join(r"C:\\", file_name)
    wb.save(dest_path)

# 파일  프레임(파일추가, 선택 삭제)
file_frame = Frame(root)
file_frame.pack(fill="x", padx=5, pady=5) # 간격 띄우기

# 불러오기 버튼
btn_add_file = Button(file_frame, text="불러오기", command=add_file, padx=5, pady=5, width=12)
btn_add_file.pack(side="left")

btn_del_file = Button(file_frame, text="선택삭제", command=del_file, padx=5, pady=5, width=12)
btn_del_file.pack(side="right")


# 파일리스트 프레임
list_frame = Frame(root)
list_frame.pack(fill="both", padx=5, pady=5)

# 파일리스트 스크롤바
scrollbar = Scrollbar(list_frame)
scrollbar.pack(side="right", fill="y")

# 파일리스트
list_file = Listbox(list_frame, selectmode="extended", height=5, yscrollcommand=scrollbar.set)
list_file.pack(side="left", fill="both", expand=True)
scrollbar.config(command=list_file.yview)


# 데이터 시작위치
opt_frame = LabelFrame(root, text="데이터시작위치")
opt_frame.pack(fill="x", padx=5, pady=5, ipady=4)

opt_rows = ["1행", "2행", "3행"]
opt_rows_combo = ttk.Combobox(opt_frame, stat="readonly", values=opt_rows)
opt_rows_combo.current(1)
opt_rows_combo.pack(side="left", padx=5, pady=5)

# 데이터 범위 프레임
range_frame = Frame(root)
range_frame.pack(fill="x")

# 데이터 범위1
range1_frame = LabelFrame(range_frame, text="데이터1 범위")
range1_frame.pack(side="left", padx=5, pady=5, ipady=4)

range1_entry = Entry(range1_frame, width=11)
range1_entry.pack(padx=5, pady=5)
range1_entry.insert(0, "A:F")

# 데이터 범위2
range2_frame = LabelFrame(range_frame, text="데이터2 범위")
range2_frame.pack(side="right", padx=5, pady=5, ipady=4)

range2_entry = Entry(range2_frame, width=11)
range2_entry.pack(padx=5, pady=5)
range2_entry.insert(0, "H:K")

# 비교 옵션1
comp1_frame = LabelFrame(root, text="비교1")
comp1_frame.pack(fill="x", padx=5, pady=5)

comp1_entry1 = Entry(comp1_frame, width=11)
comp1_entry1.pack(side="left", padx=5, pady=5)
comp1_entry1.insert(0, "B")

comp1_entry2 = Entry(comp1_frame, width=11)
comp1_entry2.pack(side="right", padx=5, pady=5)
comp1_entry2.insert(0, "I")


# 비교 옵션2
comp2_frame = LabelFrame(root, text="비교2")
comp2_frame.pack(fill="x", padx=5, pady=5)

comp2_entry1 = Entry(comp2_frame, width=11)
comp2_entry1.pack(side="left", padx=5, pady=5)
comp2_entry1.insert(0, "C")

comp2_entry2 = Entry(comp2_frame, width=11)
comp2_entry2.pack(side="right", padx=5, pady=5)
comp2_entry2.insert(0, "J")



# 실행 프레임
run_frame = Frame(root)
run_frame.pack(fill="x", padx=5, pady=5) # 간격 띄우기

# 불러오기 버튼
btn_start = Button(run_frame, text="시작", command=start, padx=5, pady=5, width=12)
btn_start.pack(side="left")

btn_close = Button(run_frame, text="닫기", command=root.quit, padx=5, pady=5, width=12)
btn_close.pack(side="right")



root.resizable(False, False)
root.mainloop()